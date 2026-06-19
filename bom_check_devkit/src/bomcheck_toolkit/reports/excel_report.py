from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from bomcheck_toolkit.models import CheckIssue, RefMatch

FILL = {
    "error": PatternFill("solid", fgColor="FFC7CE"),
    "warning": PatternFill("solid", fgColor="FFEB9C"),
    "info": PatternFill("solid", fgColor="DDEBF7"),
    "ignore": PatternFill("solid", fgColor="E7E6E6"),
}


def _format_sheet(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")
    for column_cells in ws.columns:
        max_len = 0
        col = column_cells[0].column
        for cell in column_cells:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 60)


def write_match_report(path: str | Path, matches: list[RefMatch]) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "RefDes Match"
    ws.append([
        "RefDes", "BOM_Row", "BOM_Value", "BOM_MPN", "PDF_Page",
        "PDF_Page_Name", "PDF_BBox", "PDF_Context", "Match_Count",
        "Status", "Confidence", "Note",
    ])
    for m in matches:
        bom_rows = ",".join(str(i.row_index) for i in m.bom_items)
        bom_value = " | ".join(i.value for i in m.bom_items if i.value)
        bom_mpn = " | ".join(i.mpn for i in m.bom_items if i.mpn)
        pages = ",".join(str(r.page_index + 1) for r in m.pdf_refs)
        page_name = " | ".join(dict.fromkeys(r.page_name for r in m.pdf_refs if r.page_name))
        bbox = " | ".join(",".join(f"{v:.1f}" for v in r.bbox) for r in m.pdf_refs[:2])
        context = " | ".join(r.context_text[:120] for r in m.pdf_refs[:2])
        note = "多处匹配，建议人工确认位置" if len(m.pdf_refs) > 1 else ""
        ws.append([
            m.refdes, bom_rows, bom_value, bom_mpn, pages, page_name, bbox,
            context, len(m.pdf_refs), m.status, m.confidence, note,
        ])
    _format_sheet(ws)
    wb.save(path)


def write_check_report(path: str | Path, issues: list[CheckIssue]) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Issues"
    ws.append(["Severity", "Rule_ID", "RefDes", "BOM_Row", "PDF_Page", "Title", "证据/上下文", "Suggestion", "Status"])
    rank = {"error": 0, "warning": 1, "info": 2, "ignore": 3}
    for issue in sorted(issues, key=lambda x: (rank[x.severity], x.refdes)):
        ws.append([
            issue.severity, issue.rule_id, issue.refdes, issue.bom_row, issue.pdf_page,
            issue.title, issue.evidence, issue.suggestion, issue.status,
        ])
        for cell in ws[ws.max_row]:
            cell.fill = FILL.get(issue.severity, PatternFill())
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _format_sheet(ws)
    wb.save(path)
