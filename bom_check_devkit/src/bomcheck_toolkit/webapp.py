from __future__ import annotations

import argparse
import html as html_lib
import io
import json
import mimetypes
import queue
import re
import threading
import traceback
import uuid
import zipfile
from dataclasses import dataclass
from datetime import datetime
from email import policy
from email.parser import BytesParser
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, quote, unquote, urlparse

import fitz
from openpyxl import load_workbook

from bomcheck_toolkit.pipeline import run_check

BOM_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm", ".csv"}
OUTPUT_FILES = [
    ("annotated.pdf", "标注 PDF", "pdf"),
    ("check_report.xlsx", "异常报告", "xlsx"),
    ("refdes_match_report.xlsx", "位号匹配报告", "xlsx"),
    ("bom_parsed.json", "BOM 解析 JSON", "json"),
    ("refdes_extracted.json", "PDF 位号 JSON", "json"),
]


@dataclass(slots=True)
class UploadedPart:
    filename: str
    content: bytes


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def safe_filename(filename: str | None, fallback: str) -> str:
    name = Path(filename or fallback).name.strip() or fallback
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name)
    name = re.sub(r"\s+", " ", name).strip(" .")
    if not name:
        name = fallback
    stem = Path(name).stem[:90].strip(" .") or Path(fallback).stem
    suffix = Path(name).suffix[:16]
    return f"{stem}{suffix}"


def format_size(size: int) -> str:
    units = ["B", "KB", "MB", "GB"]
    value = float(size)
    for unit in units:
        if value < 1024 or unit == units[-1]:
            return f"{value:.1f} {unit}" if unit != "B" else f"{int(value)} B"
        value /= 1024
    return f"{size} B"


def json_default(value: Any) -> str:
    return str(value)


def parse_multipart(content_type: str, body: bytes) -> dict[str, UploadedPart]:
    header = f"Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n".encode("utf-8")
    message = BytesParser(policy=policy.default).parsebytes(header + body)
    if not message.is_multipart():
        raise ValueError("请求不是 multipart/form-data")

    files: dict[str, UploadedPart] = {}
    for part in message.iter_parts():
        disposition = part.get("Content-Disposition", "")
        if "form-data" not in disposition:
            continue
        name = part.get_param("name", header="content-disposition")
        filename = part.get_filename()
        if not name or not filename:
            continue
        content = part.get_payload(decode=True) or b""
        files[name] = UploadedPart(filename=filename, content=content)
    return files


def workbook_preview(path: Path, *, max_rows: int = 80, max_cols: int = 12) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(max_row=max_rows + 1, max_col=max_cols, values_only=True):
        rows.append(["" if cell is None else str(cell)[:240] for cell in row])
    headers = rows[0] if rows else []
    data_rows = rows[1:]
    return {
        "sheet": ws.title,
        "headers": headers,
        "rows": data_rows,
        "truncated": ws.max_row > max_rows + 1 or ws.max_column > max_cols,
        "total_rows": ws.max_row,
        "total_cols": ws.max_column,
    }


class BomCheckServer(ThreadingHTTPServer):
    def __init__(
        self,
        server_address: tuple[str, int],
        jobs_root: Path,
        max_upload_bytes: int,
        worker_count: int = 1,
    ) -> None:
        super().__init__(server_address, BomCheckHandler)
        self.jobs_root = jobs_root
        self.max_upload_bytes = max_upload_bytes
        self.worker_count = max(1, worker_count)
        self.jobs: dict[str, dict[str, Any]] = {}
        self.jobs_lock = threading.Lock()
        self.job_queue: queue.Queue[tuple[str, str, str, str]] = queue.Queue()
        self.queued_job_ids: list[str] = []
        self.jobs_root.mkdir(parents=True, exist_ok=True)
        for worker_index in range(self.worker_count):
            worker = threading.Thread(
                target=self.worker_loop,
                name=f"bomcheck-worker-{worker_index + 1}",
                daemon=True,
            )
            worker.start()

    def job_public(self, job: dict[str, Any]) -> dict[str, Any]:
        hidden = {"job_dir", "input_dir", "outdir"}
        public = {key: value for key, value in job.items() if key not in hidden}
        summary = public.get("summary")
        if isinstance(summary, dict) and "files" in summary:
            summary = dict(summary)
            summary.pop("files", None)
            public["summary"] = summary
        return public

    def update_job(self, job_id: str, **updates: Any) -> None:
        with self.jobs_lock:
            job = self.jobs[job_id]
            job.update(updates)
            job["updated_at"] = now_iso()
            if job.get("status") == "queued":
                self._refresh_queue_positions_locked()

    def append_log(self, job_id: str, message: str) -> None:
        entry = f"[{datetime.now().strftime('%H:%M:%S')}] {message}"
        with self.jobs_lock:
            job = self.jobs[job_id]
            job.setdefault("log", []).append(entry)
            job["updated_at"] = now_iso()

    def get_job(self, job_id: str) -> dict[str, Any] | None:
        with self.jobs_lock:
            return self.jobs.get(job_id)

    def list_jobs(self) -> list[dict[str, Any]]:
        with self.jobs_lock:
            jobs = sorted(self.jobs.values(), key=lambda item: item["created_at"], reverse=True)
            return [self.job_public(job) for job in jobs[:30]]

    def enqueue_job(self, job_id: str, bom_path: Path, pdf_path: Path, outdir: Path) -> None:
        with self.jobs_lock:
            self.queued_job_ids.append(job_id)
            self._refresh_queue_positions_locked()
        self.job_queue.put((job_id, str(bom_path), str(pdf_path), str(outdir)))

    def worker_loop(self) -> None:
        while True:
            job_id, bom_path, pdf_path, outdir = self.job_queue.get()
            try:
                self._mark_job_dequeued(job_id)
                run_job(self, job_id, Path(bom_path), Path(pdf_path), Path(outdir))
            finally:
                self.job_queue.task_done()

    def _mark_job_dequeued(self, job_id: str) -> None:
        with self.jobs_lock:
            if job_id in self.queued_job_ids:
                self.queued_job_ids.remove(job_id)
            self._refresh_queue_positions_locked()

    def _refresh_queue_positions_locked(self) -> None:
        for index, queued_job_id in enumerate(self.queued_job_ids, start=1):
            job = self.jobs.get(queued_job_id)
            if not job or job.get("status") != "queued":
                continue
            ahead = index - 1
            job["queue_position"] = index
            job["queue_ahead"] = ahead
            job["message"] = "排队中" if ahead == 0 else f"排队中，前面还有 {ahead} 个任务"
            job["updated_at"] = now_iso()


class BomCheckHandler(BaseHTTPRequestHandler):
    server: BomCheckServer

    def do_GET(self) -> None:  # noqa: N802
        parsed = urlparse(self.path)
        path = parsed.path
        if path == "/":
            self.send_html(APP_HTML)
            return
        if path == "/health":
            with self.server.jobs_lock:
                queued = sum(1 for job in self.server.jobs.values() if job.get("status") == "queued")
                running = sum(1 for job in self.server.jobs.values() if job.get("status") == "running")
            self.send_json(
                {
                    "ok": True,
                    "time": now_iso(),
                    "queued": queued,
                    "running": running,
                    "workers": self.server.worker_count,
                }
            )
            return
        if path == "/api/jobs":
            self.send_json({"jobs": self.server.list_jobs()})
            return
        if path.startswith("/api/jobs/"):
            self.handle_job_status(path)
            return
        if path.startswith("/jobs/"):
            self.handle_job_asset(path, parse_qs(parsed.query))
            return
        self.send_error_json(HTTPStatus.NOT_FOUND, "未找到该地址")

    def do_HEAD(self) -> None:  # noqa: N802
        parsed = urlparse(self.path)
        if parsed.path.startswith("/jobs/"):
            self.handle_job_asset(parsed.path, parse_qs(parsed.query), include_body=False)
            return
        if parsed.path == "/health":
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", "0")
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            return
        self.send_response(HTTPStatus.NOT_FOUND)
        self.send_header("Content-Length", "0")
        self.end_headers()

    def do_POST(self) -> None:  # noqa: N802
        parsed = urlparse(self.path)
        if parsed.path == "/api/jobs":
            self.handle_create_job()
            return
        self.send_error_json(HTTPStatus.NOT_FOUND, "未找到该地址")

    def log_message(self, fmt: str, *args: Any) -> None:
        print(f"{self.address_string()} - {fmt % args}")

    def handle_create_job(self) -> None:
        content_length = self.headers.get("Content-Length")
        if not content_length:
            self.send_error_json(HTTPStatus.LENGTH_REQUIRED, "缺少 Content-Length")
            return
        try:
            total_bytes = int(content_length)
        except ValueError:
            self.send_error_json(HTTPStatus.BAD_REQUEST, "Content-Length 无效")
            return
        if total_bytes > self.server.max_upload_bytes:
            limit_mb = self.server.max_upload_bytes // (1024 * 1024)
            self.send_error_json(HTTPStatus.REQUEST_ENTITY_TOO_LARGE, f"上传文件超过 {limit_mb} MB 限制")
            return

        content_type = self.headers.get("Content-Type", "")
        if "multipart/form-data" not in content_type:
            self.send_error_json(HTTPStatus.BAD_REQUEST, "请使用 multipart/form-data 上传")
            return

        try:
            body = self.rfile.read(total_bytes)
            files = parse_multipart(content_type, body)
            bom = files.get("bom")
            pdf = files.get("pdf")
            if not bom or not bom.content:
                raise ValueError("缺少 BOM Excel")
            if not pdf or not pdf.content:
                raise ValueError("缺少原理图 PDF")

            bom_name = safe_filename(bom.filename, "bom.xlsx")
            pdf_name = safe_filename(pdf.filename, "schematic.pdf")
            bom_ext = Path(bom_name).suffix.lower()
            pdf_ext = Path(pdf_name).suffix.lower()
            if bom_ext not in BOM_EXTENSIONS:
                raise ValueError("BOM 仅支持 xlsx/xlsm/xltx/xltm/csv")
            if pdf_ext != ".pdf":
                raise ValueError("原理图仅支持 PDF")

            job_id = f"{datetime.now().strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:8]}"
            job_dir = self.server.jobs_root / job_id
            input_dir = job_dir / "input"
            outdir = job_dir / "output"
            input_dir.mkdir(parents=True, exist_ok=True)
            outdir.mkdir(parents=True, exist_ok=True)

            bom_path = input_dir / bom_name
            pdf_path = input_dir / pdf_name
            bom_path.write_bytes(bom.content)
            pdf_path.write_bytes(pdf.content)

            job = {
                "id": job_id,
                "status": "queued",
                "message": "已加入队列",
                "created_at": now_iso(),
                "updated_at": now_iso(),
                "log": [],
                "inputs": {
                    "bom": bom_name,
                    "pdf": pdf_name,
                    "bom_size": format_size(len(bom.content)),
                    "pdf_size": format_size(len(pdf.content)),
                },
                "summary": {},
                "files": [],
                "previews": {},
                "queue_position": None,
                "queue_ahead": None,
                "job_dir": job_dir,
                "input_dir": input_dir,
                "outdir": outdir,
            }
            with self.server.jobs_lock:
                self.server.jobs[job_id] = job
            self.server.append_log(job_id, "收到上传文件")
            self.server.enqueue_job(job_id, bom_path, pdf_path, outdir)
            self.send_json({"job": self.server.job_public(job)}, status=HTTPStatus.ACCEPTED)
        except ValueError as exc:
            self.send_error_json(HTTPStatus.BAD_REQUEST, str(exc))
        except Exception as exc:  # pragma: no cover - defensive request guard
            self.send_error_json(HTTPStatus.INTERNAL_SERVER_ERROR, f"创建任务失败: {exc}")

    def handle_job_status(self, path: str) -> None:
        parts = path.strip("/").split("/")
        if len(parts) != 3:
            self.send_error_json(HTTPStatus.NOT_FOUND, "任务地址无效")
            return
        job = self.server.get_job(parts[2])
        if not job:
            self.send_error_json(HTTPStatus.NOT_FOUND, "任务不存在")
            return
        self.send_json({"job": self.server.job_public(job)})

    def handle_job_asset(self, path: str, query: dict[str, list[str]], include_body: bool = True) -> None:
        parts = path.strip("/").split("/")
        if len(parts) == 3:
            job_id = parts[1]
            action = parts[2]
            if action == "locate.png":
                self.handle_location_image(job_id, query, include_body=include_body)
                return
            if action == "detail":
                self.handle_detail_page(job_id, include_body=include_body)
                return
            if action == "report.html":
                self.handle_report_page(job_id, query, include_body=include_body)
                return
            if action == "bundle.zip":
                self.handle_bundle_zip(job_id, include_body=include_body)
                return
            if action == "excel.zip":
                self.handle_excel_zip(job_id, include_body=include_body)
                return
            self.send_error_json(HTTPStatus.NOT_FOUND, "文件地址无效")
            return
        if len(parts) != 4 or parts[2] != "files":
            self.send_error_json(HTTPStatus.NOT_FOUND, "文件地址无效")
            return
        job_id = parts[1]
        filename = safe_filename(unquote(parts[3]), "output.bin")
        job = self.server.get_job(job_id)
        if not job:
            self.send_error_json(HTTPStatus.NOT_FOUND, "任务不存在")
            return

        outdir = Path(job["outdir"]).resolve()
        file_path = (outdir / filename).resolve()
        if not file_path.is_file() or not file_path.is_relative_to(outdir):
            self.send_error_json(HTTPStatus.NOT_FOUND, "文件不存在")
            return

        content = file_path.read_bytes()
        mime = mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"
        download = query.get("download", ["0"])[0] == "1"
        disposition = "attachment" if download else "inline"
        encoded_name = quote(file_path.name)
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", mime)
        self.send_header("Content-Length", str(len(content)))
        self.send_header("Content-Disposition", f"{disposition}; filename*=UTF-8''{encoded_name}")
        self.send_header("Cache-Control", "private, max-age=3600")
        self.end_headers()
        if include_body:
            self.wfile.write(content)

    def handle_detail_page(self, job_id: str, include_body: bool = True) -> None:
        job = self.server.get_job(job_id)
        if not job:
            self.send_error_json(HTTPStatus.NOT_FOUND, "任务不存在")
            return
        self.send_html(build_detail_html(self.server.job_public(job)), include_body=include_body)

    def handle_report_page(
        self,
        job_id: str,
        query: dict[str, list[str]],
        include_body: bool = True,
    ) -> None:
        job = self.server.get_job(job_id)
        if not job:
            self.send_error_json(HTTPStatus.NOT_FOUND, "任务不存在")
            return
        content = build_report_html(self.server.job_public(job)).encode("utf-8")
        download = query.get("download", ["0"])[0] == "1"
        disposition = "attachment" if download else "inline"
        self.send_binary(
            content,
            "text/html; charset=utf-8",
            filename=f"bom_check_{job_id}_report.html",
            disposition=disposition,
            cache_control="no-store",
            include_body=include_body,
        )

    def handle_bundle_zip(self, job_id: str, include_body: bool = True) -> None:
        job = self.server.get_job(job_id)
        if not job:
            self.send_error_json(HTTPStatus.NOT_FOUND, "任务不存在")
            return
        if job.get("status") != "done":
            self.send_error_json(HTTPStatus.CONFLICT, "任务尚未完成，不能打包下载")
            return
        try:
            content = build_bundle_zip(self.server.job_public(job), Path(job["outdir"]))
        except Exception as exc:  # pragma: no cover - defensive archive guard
            self.send_error_json(HTTPStatus.INTERNAL_SERVER_ERROR, f"生成打包文件失败: {exc}")
            return
        self.send_binary(
            content,
            "application/zip",
            filename=f"bom_check_{job_id}_bundle.zip",
            disposition="attachment",
            cache_control="private, max-age=300",
            include_body=include_body,
        )

    def handle_excel_zip(self, job_id: str, include_body: bool = True) -> None:
        job = self.server.get_job(job_id)
        if not job:
            self.send_error_json(HTTPStatus.NOT_FOUND, "任务不存在")
            return
        if job.get("status") != "done":
            self.send_error_json(HTTPStatus.CONFLICT, "任务尚未完成，不能下载 Excel")
            return
        try:
            content = build_excel_zip(Path(job["outdir"]))
        except FileNotFoundError:
            self.send_error_json(HTTPStatus.NOT_FOUND, "Excel 输出不存在")
            return
        except Exception as exc:  # pragma: no cover - defensive archive guard
            self.send_error_json(HTTPStatus.INTERNAL_SERVER_ERROR, f"生成 Excel 包失败: {exc}")
            return
        self.send_binary(
            content,
            "application/zip",
            filename=f"bom_check_{job_id}_excel.zip",
            disposition="attachment",
            cache_control="private, max-age=300",
            include_body=include_body,
        )

    def handle_location_image(self, job_id: str, query: dict[str, list[str]], include_body: bool = True) -> None:
        job = self.server.get_job(job_id)
        if not job:
            self.send_error_json(HTTPStatus.NOT_FOUND, "任务不存在")
            return
        try:
            page_index = int(query.get("page_index", [""])[0])
            bbox = [float(value) for value in query.get("bbox", [""])[0].split(",")]
            if len(bbox) != 4:
                raise ValueError
        except (TypeError, ValueError):
            self.send_error_json(HTTPStatus.BAD_REQUEST, "定位参数无效")
            return

        pdf_path = Path(job["outdir"]) / "annotated.pdf"
        if not pdf_path.is_file():
            self.send_error_json(HTTPStatus.NOT_FOUND, "标注 PDF 不存在")
            return

        try:
            with fitz.open(pdf_path) as doc:
                if page_index < 0 or page_index >= len(doc):
                    self.send_error_json(HTTPStatus.BAD_REQUEST, "页码超出范围")
                    return
                page = doc[page_index]
                rect = fitz.Rect(bbox) + (-5, -5, 5, 5)
                page.draw_rect(rect, color=(0.0, 0.28, 1.0), width=4.5, overlay=True)
                page.draw_rect(rect + (-4, -4, 4, 4), color=(1.0, 0.82, 0.0), width=2.2, overlay=True)
                pix = page.get_pixmap(matrix=fitz.Matrix(1.7, 1.7), alpha=False)
                content = pix.tobytes("png")
        except Exception as exc:  # pragma: no cover - defensive render guard
            self.send_error_json(HTTPStatus.INTERNAL_SERVER_ERROR, f"生成定位图失败: {exc}")
            return

        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "image/png")
        self.send_header("Content-Length", str(len(content)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        if include_body:
            self.wfile.write(content)

    def send_binary(
        self,
        content: bytes,
        mime: str,
        *,
        filename: str | None = None,
        disposition: str = "inline",
        cache_control: str = "no-store",
        include_body: bool = True,
        status: HTTPStatus = HTTPStatus.OK,
    ) -> None:
        self.send_response(status)
        self.send_header("Content-Type", mime)
        self.send_header("Content-Length", str(len(content)))
        if filename:
            encoded_name = quote(filename)
            self.send_header("Content-Disposition", f"{disposition}; filename*=UTF-8''{encoded_name}")
        self.send_header("Cache-Control", cache_control)
        self.end_headers()
        if include_body:
            self.wfile.write(content)

    def send_html(
        self,
        html: str,
        status: HTTPStatus = HTTPStatus.OK,
        *,
        include_body: bool = True,
    ) -> None:
        content = html.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(content)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        if include_body:
            self.wfile.write(content)

    def send_json(self, data: dict[str, Any], status: HTTPStatus = HTTPStatus.OK) -> None:
        content = json.dumps(data, ensure_ascii=False, default=json_default).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(content)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(content)

    def send_error_json(self, status: HTTPStatus, message: str) -> None:
        self.send_json({"error": message}, status=status)


def run_job(server: BomCheckServer, job_id: str, bom_path: Path, pdf_path: Path, outdir: Path) -> None:
    server.update_job(job_id, status="running", message="检查中", queue_position=None, queue_ahead=None)
    try:
        summary = run_check(
            bom_path,
            pdf_path,
            outdir,
            logger=lambda message: server.append_log(job_id, message),
        )
        files = build_file_list(job_id, outdir)
        issues_preview = workbook_preview(outdir / "check_report.xlsx")
        issues_preview["targets"] = summary.get("issue_targets", [])[: len(issues_preview.get("rows", []))]
        previews = {
            "issues": issues_preview,
            "matches": workbook_preview(outdir / "refdes_match_report.xlsx"),
        }
        server.update_job(
            job_id,
            status="done",
            message="完成",
            summary=summary,
            files=files,
            previews=previews,
        )
    except Exception as exc:
        server.append_log(job_id, traceback.format_exc())
        server.update_job(job_id, status="failed", message=str(exc))


def build_file_list(job_id: str, outdir: Path) -> list[dict[str, Any]]:
    files: list[dict[str, Any]] = []
    for filename, label, kind in OUTPUT_FILES:
        path = outdir / filename
        if not path.exists():
            continue
        url = f"/jobs/{job_id}/files/{quote(filename)}"
        files.append(
            {
                "name": filename,
                "label": label,
                "kind": kind,
                "size": format_size(path.stat().st_size),
                "url": url,
                "download_url": f"{url}?download=1",
            }
        )
    return files


APP_HTML = r"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>BOM Check</title>
  <style>
    :root {
      color-scheme: light;
      --bg: #f6f7f9;
      --panel: #ffffff;
      --panel-soft: #f0f7f5;
      --ink: #16211f;
      --muted: #60706c;
      --line: #d8e0dd;
      --primary: #0b6b5f;
      --primary-strong: #074d44;
      --accent: #e97b35;
      --danger: #b42318;
      --warn: #b65c13;
      --ok: #11845b;
      --focus: #2563eb;
      --radius: 8px;
      font-family: Inter, "Segoe UI", "Microsoft YaHei", system-ui, sans-serif;
    }

    * {
      box-sizing: border-box;
    }

    body {
      margin: 0;
      min-width: 320px;
      background: var(--bg);
      color: var(--ink);
      line-height: 1.5;
    }

    button,
    input {
      font: inherit;
    }

    button,
    .file-name,
    .download-link,
    .tab {
      transition: background-color 180ms ease, border-color 180ms ease, color 180ms ease;
    }

    button:focus-visible,
    input:focus-visible,
    a:focus-visible {
      outline: 3px solid color-mix(in srgb, var(--focus) 45%, transparent);
      outline-offset: 2px;
    }

    .app {
      width: min(1500px, calc(100vw - 32px));
      margin: 0 auto;
      padding: 24px 0 32px;
    }

    .topbar {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 16px;
      margin-bottom: 16px;
    }

    h1 {
      margin: 0;
      font-size: 28px;
      line-height: 1.15;
      letter-spacing: 0;
    }

    .subtitle {
      margin: 4px 0 0;
      color: var(--muted);
      font-size: 14px;
    }

    .status-pill {
      min-height: 34px;
      padding: 6px 12px;
      border: 1px solid var(--line);
      border-radius: 999px;
      background: var(--panel);
      color: var(--muted);
      font-weight: 600;
      white-space: nowrap;
    }

    .status-pill.running {
      color: var(--warn);
      border-color: #f0c69c;
      background: #fff8ef;
    }

    .status-pill.done {
      color: var(--ok);
      border-color: #a7d8c4;
      background: #effaf5;
    }

    .status-pill.failed {
      color: var(--danger);
      border-color: #f0b8b2;
      background: #fff3f1;
    }

    .top-actions {
      display: flex;
      align-items: center;
      gap: 10px;
    }

    .top-link {
      min-height: 34px;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      padding: 6px 12px;
      border: 1px solid var(--line);
      border-radius: 6px;
      color: var(--primary);
      background: #fff;
      font-size: 13px;
      font-weight: 700;
      text-decoration: none;
    }

    .top-link:hover {
      border-color: var(--primary);
      background: var(--panel-soft);
    }

    .grid {
      display: grid;
      grid-template-columns: minmax(320px, 410px) minmax(0, 1fr);
      gap: 16px;
      align-items: start;
    }

    .surface {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: var(--radius);
    }

    .upload {
      padding: 16px;
    }

    .field {
      display: grid;
      gap: 8px;
      margin-bottom: 14px;
    }

    label {
      color: var(--muted);
      font-size: 13px;
      font-weight: 700;
    }

    input[type="file"] {
      width: 100%;
      min-height: 44px;
      padding: 10px;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: #fff;
      color: var(--ink);
      cursor: pointer;
    }

    .actions {
      display: flex;
      gap: 10px;
      align-items: center;
      margin-top: 4px;
    }

    .primary-btn,
    .secondary-btn,
    .tab {
      min-height: 44px;
      border-radius: 6px;
      border: 1px solid transparent;
      cursor: pointer;
      font-weight: 700;
    }

    .primary-btn {
      flex: 1;
      color: #fff;
      background: var(--primary);
      border-color: var(--primary);
    }

    .primary-btn:hover {
      background: var(--primary-strong);
    }

    .primary-btn:disabled {
      cursor: not-allowed;
      background: #98aaa5;
      border-color: #98aaa5;
    }

    .secondary-btn {
      padding: 0 14px;
      color: var(--primary);
      background: #fff;
      border-color: var(--line);
    }

    .secondary-btn:hover {
      border-color: var(--primary);
      background: var(--panel-soft);
    }

    .upload-progress {
      height: 8px;
      margin-top: 14px;
      overflow: hidden;
      border-radius: 999px;
      background: #e7ecea;
    }

    .upload-progress span {
      display: block;
      width: 0;
      height: 100%;
      background: var(--primary);
    }

    .metrics {
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 8px;
      padding: 12px;
      border-bottom: 1px solid var(--line);
    }

    .metric {
      min-height: 72px;
      padding: 10px;
      border: 1px solid #e3ebe8;
      border-radius: 6px;
      background: #fbfcfc;
    }

    .metric .value {
      display: block;
      font-size: 24px;
      line-height: 1.1;
      font-weight: 800;
    }

    .metric .label {
      display: block;
      margin-top: 4px;
      color: var(--muted);
      font-size: 12px;
      font-weight: 700;
    }

    .result-body {
      padding: 12px;
    }

    .downloads {
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(170px, 1fr));
      gap: 8px;
      margin-bottom: 12px;
    }

    .download-link {
      min-height: 44px;
      padding: 9px 10px;
      border: 1px solid var(--line);
      border-radius: 6px;
      color: var(--ink);
      text-decoration: none;
      background: #fff;
    }

    .download-link:hover {
      border-color: var(--primary);
      background: var(--panel-soft);
    }

    .download-link strong,
    .download-link span {
      display: block;
      overflow-wrap: anywhere;
    }

    .download-link span {
      color: var(--muted);
      font-size: 12px;
    }

    .preview-layout {
      display: grid;
      grid-template-columns: minmax(520px, 1fr) minmax(520px, 1fr);
      gap: 12px;
      align-items: start;
    }

    iframe {
      width: 100%;
      height: min(70vh, 760px);
      min-height: 560px;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: #fff;
    }

    .preview-pane {
      min-width: 0;
    }

    .pane-toolbar {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 10px;
      margin-bottom: 8px;
    }

    .pane-title {
      margin: 0;
      color: var(--muted);
      font-size: 13px;
      font-weight: 800;
    }

    .locator {
      margin-top: 10px;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: #fff;
      overflow: hidden;
    }

    .locator-head {
      padding: 8px 10px;
      border-bottom: 1px solid var(--line);
      background: #fff9e7;
      color: #3b3121;
      font-size: 13px;
      font-weight: 800;
    }

    .locator img {
      display: block;
      width: 100%;
      max-height: min(70vh, 760px);
      object-fit: contain;
      background: #fff;
    }

    .tabs {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin-bottom: 8px;
    }

    .tab {
      padding: 0 12px;
      color: var(--muted);
      background: #fff;
      border-color: var(--line);
    }

    .tab.active {
      color: #fff;
      background: var(--primary);
      border-color: var(--primary);
    }

    .table-wrap {
      overflow: auto;
      max-height: 430px;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: #fff;
    }

    table {
      width: 100%;
      border-collapse: collapse;
      min-width: 760px;
      font-size: 13px;
    }

    th,
    td {
      padding: 8px 10px;
      border-bottom: 1px solid #e7ecea;
      text-align: left;
      vertical-align: top;
    }

    th {
      position: sticky;
      top: 0;
      z-index: 1;
      color: #20312e;
      background: #eaf3f0;
      font-size: 12px;
    }

    td {
      max-width: 280px;
      overflow-wrap: anywhere;
    }

    tr.locatable {
      cursor: pointer;
    }

    tr.locatable:hover td {
      background: #f0f7f5;
    }

    tr.selected td {
      background: #fff3bd;
    }

    tr.no-location td:first-child::after,
    tr.locatable td:first-child::after {
      display: block;
      margin-top: 4px;
      color: var(--muted);
      font-size: 11px;
      font-weight: 700;
    }

    tr.locatable td:first-child::after {
      content: "点击定位";
      color: var(--primary);
    }

    tr.no-location td:first-child::after {
      content: "无坐标";
    }

    .log {
      margin-top: 16px;
      padding: 12px;
      color: #23312f;
      background: #f9fbfb;
      border-top: 1px solid var(--line);
      font-family: "SFMono-Regular", Consolas, "Liberation Mono", monospace;
      font-size: 12px;
      white-space: pre-wrap;
      max-height: 220px;
      overflow: auto;
    }

    .empty {
      min-height: 220px;
      display: grid;
      place-items: center;
      padding: 24px;
      color: var(--muted);
      text-align: center;
    }

    .error-text {
      color: var(--danger);
      font-weight: 700;
    }

    @media (max-width: 1180px) {
      .grid,
      .metrics,
      .preview-layout {
        grid-template-columns: 1fr;
      }

      iframe {
        min-height: 420px;
      }
    }

    @media (prefers-reduced-motion: reduce) {
      *,
      *::before,
      *::after {
        transition-duration: 1ms !important;
        scroll-behavior: auto !important;
      }
    }
  </style>
</head>
<body>
  <main class="app">
    <header class="topbar">
      <div>
        <h1>BOM Check</h1>
        <p class="subtitle">Excel + 原理图 PDF</p>
      </div>
      <div class="top-actions">
        <a class="top-link" href="/" target="_blank" rel="noopener">独立窗口打开</a>
        <div id="statusPill" class="status-pill">就绪</div>
      </div>
    </header>

    <div class="grid">
      <form id="uploadForm" class="surface upload">
        <div class="field">
          <label for="bom">BOM Excel</label>
          <input id="bom" name="bom" type="file" accept=".xlsx,.xlsm,.xltx,.xltm,.csv" required>
        </div>
        <div class="field">
          <label for="pdf">原理图 PDF</label>
          <input id="pdf" name="pdf" type="file" accept=".pdf" required>
        </div>
        <div class="actions">
          <button id="runButton" class="primary-btn" type="submit">运行检查</button>
          <button id="refreshButton" class="secondary-btn" type="button">刷新</button>
        </div>
        <div class="upload-progress" aria-hidden="true"><span id="progressBar"></span></div>
        <div id="formError" class="error-text" role="alert"></div>
      </form>

      <section id="resultPanel" class="surface" aria-live="polite">
        <div class="empty">暂无任务</div>
      </section>
    </div>
  </main>

  <script>
    const form = document.getElementById("uploadForm");
    const runButton = document.getElementById("runButton");
    const refreshButton = document.getElementById("refreshButton");
    const statusPill = document.getElementById("statusPill");
    const resultPanel = document.getElementById("resultPanel");
    const progressBar = document.getElementById("progressBar");
    const formError = document.getElementById("formError");

    let currentJobId = null;
    let pollTimer = null;
    let activePreview = "issues";
    let selectedTarget = null;

    function escapeHtml(value) {
      return String(value ?? "")
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;")
        .replaceAll("'", "&#039;");
    }

    function setStatus(text, state) {
      statusPill.textContent = text;
      statusPill.className = "status-pill" + (state ? " " + state : "");
    }

    function metric(label, value) {
      return `<div class="metric"><span class="value">${escapeHtml(value ?? 0)}</span><span class="label">${escapeHtml(label)}</span></div>`;
    }

    function renderTable(preview, kind, jobId) {
      if (!preview || !preview.headers || preview.headers.length === 0) {
        return `<div class="empty">暂无预览</div>`;
      }
      const headers = preview.headers.map((cell) => `<th>${escapeHtml(cell)}</th>`).join("");
      const targets = preview.targets || [];
      const rows = preview.rows.map((row, rowIndex) => {
        const target = targets[rowIndex];
        const isIssueTable = kind === "issues";
        const canLocate = isIssueTable && target && target.has_location;
        const selected = selectedTarget && selectedTarget.jobId === jobId && selectedTarget.rowIndex === rowIndex;
        const rowClass = [
          canLocate ? "locatable" : "",
          isIssueTable && target && !target.has_location ? "no-location" : "",
          selected ? "selected" : "",
        ].filter(Boolean).join(" ");
        const targetAttr = isIssueTable && target
          ? ` data-target-index="${rowIndex}"`
          : "";
        const cells = preview.headers.map((_, index) => `<td>${escapeHtml(row[index] ?? "")}</td>`).join("");
        return `<tr class="${rowClass}"${targetAttr}>${cells}</tr>`;
      }).join("");
      const note = preview.truncated
        ? `<p class="subtitle">已显示前 ${preview.rows.length} 行 / 共 ${preview.total_rows} 行</p>`
        : "";
      return `${note}<div class="table-wrap"><table><thead><tr>${headers}</tr></thead><tbody>${rows}</tbody></table></div>`;
    }

    function pdfSrc(pdf, target) {
      if (!pdf) return "";
      if (target && target.has_location && target.page) {
        return `${pdf.url}#page=${target.page}&view=FitH`;
      }
      return `${pdf.url}#view=FitH`;
    }

    function renderLocator(job, pdf) {
      if (!pdf) {
        return `<div class="empty">暂无 PDF</div>`;
      }
      if (!selectedTarget || selectedTarget.jobId !== job.id) {
        return `<div class="locator"><div class="empty">点击异常报告中的可定位行</div></div>`;
      }
      if (!selectedTarget.has_location) {
        return `<div class="locator"><div class="empty">该异常没有 PDF 坐标，通常是 BOM-only 或纯表格类问题</div></div>`;
      }
      const bbox = encodeURIComponent(selectedTarget.bbox.join(","));
      const img = `/jobs/${encodeURIComponent(job.id)}/locate.png?page_index=${selectedTarget.page_index}&bbox=${bbox}&t=${Date.now()}`;
      const pdfPageUrl = `${pdf.url}#page=${selectedTarget.page}&view=FitH`;
      return `
        <div class="locator">
          <div class="locator-head">
            ${escapeHtml(selectedTarget.refdes || "")}
            · 第 ${escapeHtml(selectedTarget.page)} 页
            · ${escapeHtml(selectedTarget.title || selectedTarget.rule_id || "")}
            · <a href="${pdfPageUrl}" target="_blank" rel="noopener">新窗口打开该页</a>
          </div>
          <img src="${img}" alt="异常定位高亮预览">
        </div>
      `;
    }

    function renderJob(job) {
      if (selectedTarget && selectedTarget.jobId !== job.id) {
        selectedTarget = null;
      }
      const summary = job.summary || {};
      const counts = summary.issue_counts || {};
      const files = job.files || [];
      const pdf = files.find((file) => file.kind === "pdf");
      const downloads = files.map((file) => `
        <a class="download-link" href="${file.download_url}">
          <strong>${escapeHtml(file.label)}</strong>
          <span>${escapeHtml(file.name)} · ${escapeHtml(file.size)}</span>
        </a>
      `).join("");
      const preview = job.previews ? job.previews[activePreview] : null;
      const log = (job.log || []).map(escapeHtml).join("\n");

      resultPanel.innerHTML = `
        <div class="metrics">
          ${metric("BOM 行", summary.bom_items)}
          ${metric("PDF 位号", summary.pdf_refs)}
          ${metric("错误", counts.error || 0)}
          ${metric("警告", counts.warning || 0)}
        </div>
        <div class="result-body">
          ${downloads ? `<div class="downloads">${downloads}</div>` : ""}
          ${job.status === "failed" ? `<p class="error-text">${escapeHtml(job.message)}</p>` : ""}
          ${job.status === "done" ? `
            <div class="preview-layout">
              <div class="preview-pane">
                <div class="pane-toolbar">
                  <p class="pane-title">PDF 预览</p>
                  ${pdf ? `<a class="top-link" href="${pdf.url}" target="_blank" rel="noopener">完整 PDF 新窗口</a>` : ""}
                </div>
                ${pdf ? `<iframe id="pdfFrame" title="标注 PDF 预览" src="${pdfSrc(pdf, selectedTarget)}"></iframe>` : ""}
                ${renderLocator(job, pdf)}
              </div>
              <div class="preview-pane">
                <div class="tabs">
                  <button type="button" class="tab ${activePreview === "issues" ? "active" : ""}" data-preview="issues">异常报告</button>
                  <button type="button" class="tab ${activePreview === "matches" ? "active" : ""}" data-preview="matches">位号匹配</button>
                </div>
                <div id="tablePreview">${renderTable(preview, activePreview, job.id)}</div>
              </div>
            </div>
          ` : `<div class="empty">${escapeHtml(job.message || "处理中")}</div>`}
        </div>
        <pre class="log">${log}</pre>
      `;

      resultPanel.querySelectorAll("[data-preview]").forEach((button) => {
        button.addEventListener("click", () => {
          activePreview = button.dataset.preview;
          renderJob(job);
        });
      });

      resultPanel.querySelectorAll("[data-target-index]").forEach((row) => {
        row.addEventListener("click", () => {
          const preview = job.previews ? job.previews.issues : null;
          const rowIndex = Number(row.dataset.targetIndex);
          const target = preview && preview.targets ? preview.targets[rowIndex] : null;
          if (!target) return;
          selectedTarget = { ...target, jobId: job.id, rowIndex };
          activePreview = "issues";
          renderJob(job);
        });
      });

      if (job.status === "done") {
        setStatus("完成", "done");
      } else if (job.status === "failed") {
        setStatus("失败", "failed");
      } else if (job.status === "running" || job.status === "queued") {
        setStatus(job.status === "running" ? "运行中" : (job.message || "排队中"), "running");
      } else {
        setStatus("就绪", "");
      }
    }

    async function pollJob(jobId) {
      const response = await fetch(`/api/jobs/${encodeURIComponent(jobId)}`, { cache: "no-store" });
      const data = await response.json();
      if (!response.ok) {
        throw new Error(data.error || "查询任务失败");
      }
      renderJob(data.job);
      if (data.job.status === "done" || data.job.status === "failed") {
        runButton.disabled = false;
        progressBar.style.width = "0";
        clearTimeout(pollTimer);
        pollTimer = null;
        return;
      }
      pollTimer = setTimeout(() => pollJob(jobId).catch(showError), 1200);
    }

    function showError(error) {
      formError.textContent = error.message || String(error);
      setStatus("失败", "failed");
      runButton.disabled = false;
      progressBar.style.width = "0";
    }

    form.addEventListener("submit", (event) => {
      event.preventDefault();
      formError.textContent = "";
      runButton.disabled = true;
      setStatus("上传中", "running");
      progressBar.style.width = "0";

      const data = new FormData(form);
      const xhr = new XMLHttpRequest();
      xhr.open("POST", "/api/jobs");
      xhr.upload.onprogress = (event) => {
        if (event.lengthComputable) {
          progressBar.style.width = `${Math.round((event.loaded / event.total) * 100)}%`;
        }
      };
      xhr.onload = () => {
        try {
          const response = JSON.parse(xhr.responseText);
          if (xhr.status < 200 || xhr.status >= 300) {
            throw new Error(response.error || "上传失败");
          }
          currentJobId = response.job.id;
          activePreview = "issues";
          selectedTarget = null;
          renderJob(response.job);
          pollJob(currentJobId).catch(showError);
        } catch (error) {
          showError(error);
        }
      };
      xhr.onerror = () => showError(new Error("网络连接失败"));
      xhr.send(data);
    });

    refreshButton.addEventListener("click", () => {
      if (currentJobId) {
        pollJob(currentJobId).catch(showError);
      }
    });
  </script>
</body>
</html>
"""


def html_escape(value: Any) -> str:
    return html_lib.escape("" if value is None else str(value), quote=True)


def json_for_script(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False, default=json_default).replace("</", "<\\/")


def find_output_file(job: dict[str, Any], *, name: str | None = None, kind: str | None = None) -> dict[str, Any] | None:
    for file_info in job.get("files", []):
        if name is not None and file_info.get("name") == name:
            return file_info
        if kind is not None and file_info.get("kind") == kind:
            return file_info
    return None


def build_report_html(job: dict[str, Any]) -> str:
    job_id = str(job.get("id", ""))
    summary = job.get("summary") or {}
    issue_counts = summary.get("issue_counts") or {}
    inputs = job.get("inputs") or {}
    issues = (job.get("previews") or {}).get("issues") or {}
    headers = issues.get("headers") or []
    rows = issues.get("rows") or []
    targets = issues.get("targets") or []
    files = job.get("files") or []
    pdf_file = find_output_file(job, kind="pdf")

    action_links = [
        (f"/jobs/{quote(job_id)}/detail", "查看明细"),
        (f"/jobs/{quote(job_id)}/bundle.zip", "打包下载"),
        (f"/jobs/{quote(job_id)}/excel.zip", "下载 Excel"),
    ]
    if pdf_file:
        action_links.append((pdf_file.get("download_url", ""), "下载 PDF"))
    action_links.extend((file_info.get("download_url", ""), file_info.get("label", file_info.get("name", ""))) for file_info in files)
    actions_html = "".join(
        f'<a href="{html_escape(url)}">{html_escape(label)}</a>'
        for url, label in action_links
        if url and label
    )

    metrics = [
        ("BOM 行", summary.get("bom_items", 0)),
        ("PDF 位号", summary.get("pdf_refs", 0)),
        ("异常", summary.get("issues", 0)),
        ("错误", issue_counts.get("error", 0)),
        ("警告", issue_counts.get("warning", 0)),
    ]
    metrics_html = "".join(
        f'<div class="metric"><strong>{html_escape(value)}</strong><span>{html_escape(label)}</span></div>'
        for label, value in metrics
    )

    if headers and rows:
        header_html = "".join(f"<th>{html_escape(header)}</th>" for header in headers)
        row_html_parts = []
        for index, row in enumerate(rows):
            target = targets[index] if index < len(targets) else {}
            locate = ""
            if target.get("has_location"):
                locate = f"第 {target.get('page')} 页"
            elif target:
                locate = "无坐标"
            cells_html = "".join(
                f"<td>{html_escape(row[col_index] if col_index < len(row) else '')}</td>"
                for col_index in range(len(headers))
            )
            row_html_parts.append(
                f'<tr><td class="locate">{html_escape(locate)}</td>{cells_html}</tr>'
            )
        issue_table = (
            '<div class="table-wrap"><table><thead><tr><th>定位</th>'
            f"{header_html}</tr></thead><tbody>{''.join(row_html_parts)}</tbody></table></div>"
        )
    else:
        issue_table = '<div class="empty">暂无异常明细</div>'

    log_html = "\n".join(html_escape(item) for item in job.get("log", []))
    generated_at = now_iso()
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>BOM Check 报告 {html_escape(job_id)}</title>
  <style>
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: #f6f7f9;
      color: #17211f;
      font-family: "Segoe UI", "Microsoft YaHei", Arial, sans-serif;
      line-height: 1.55;
    }}
    main {{ width: min(1280px, calc(100vw - 40px)); margin: 0 auto; padding: 28px 0 40px; }}
    header {{ display: flex; justify-content: space-between; gap: 16px; align-items: flex-start; margin-bottom: 18px; }}
    h1 {{ margin: 0; font-size: 28px; letter-spacing: 0; }}
    .meta {{ color: #60706c; font-size: 13px; }}
    .actions {{ display: flex; flex-wrap: wrap; gap: 8px; justify-content: flex-end; }}
    .actions a {{ min-height: 36px; padding: 7px 12px; border: 1px solid #cfd9d5; border-radius: 6px; color: #0b6b5f; background: #fff; text-decoration: none; font-weight: 700; }}
    .panel {{ margin-top: 14px; border: 1px solid #d8e0dd; border-radius: 8px; background: #fff; overflow: hidden; }}
    .panel h2 {{ margin: 0; padding: 12px 14px; border-bottom: 1px solid #d8e0dd; font-size: 16px; }}
    .metrics {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); gap: 10px; padding: 14px; }}
    .metric {{ border: 1px solid #e1e8e5; border-radius: 6px; padding: 10px; background: #fbfcfc; }}
    .metric strong {{ display: block; font-size: 24px; line-height: 1.1; }}
    .metric span {{ display: block; margin-top: 4px; color: #60706c; font-size: 12px; font-weight: 700; }}
    .kv {{ display: grid; grid-template-columns: 120px 1fr; gap: 8px 12px; padding: 14px; font-size: 14px; }}
    .kv dt {{ color: #60706c; font-weight: 700; }}
    .kv dd {{ margin: 0; overflow-wrap: anywhere; }}
    .table-wrap {{ overflow: auto; }}
    table {{ width: 100%; min-width: 980px; border-collapse: collapse; font-size: 13px; }}
    th, td {{ padding: 8px 10px; border-bottom: 1px solid #e6ece9; text-align: left; vertical-align: top; overflow-wrap: anywhere; }}
    th {{ position: sticky; top: 0; background: #eaf3f0; color: #23312f; }}
    .locate {{ color: #0b6b5f; font-weight: 700; white-space: nowrap; }}
    pre {{ margin: 0; padding: 14px; max-height: 260px; overflow: auto; background: #fbfcfc; font-size: 12px; }}
    .empty {{ padding: 18px; color: #60706c; }}
    @media print {{ body {{ background: #fff; }} main {{ width: auto; padding: 0; }} .actions {{ display: none; }} }}
  </style>
</head>
<body>
  <main>
    <header>
      <div>
        <h1>BOM Check 报告</h1>
        <div class="meta">任务 {html_escape(job_id)} · 生成 {html_escape(generated_at)}</div>
      </div>
      <nav class="actions">{actions_html}</nav>
    </header>
    <section class="panel">
      <h2>结果概览</h2>
      <div class="metrics">{metrics_html}</div>
    </section>
    <section class="panel">
      <h2>输入文件</h2>
      <dl class="kv">
        <dt>BOM</dt><dd>{html_escape(inputs.get("bom", ""))} · {html_escape(inputs.get("bom_size", ""))}</dd>
        <dt>PDF</dt><dd>{html_escape(inputs.get("pdf", ""))} · {html_escape(inputs.get("pdf_size", ""))}</dd>
        <dt>状态</dt><dd>{html_escape(job.get("message", job.get("status", "")))}</dd>
      </dl>
    </section>
    <section class="panel">
      <h2>异常明细</h2>
      {issue_table}
    </section>
    <section class="panel">
      <h2>运行日志</h2>
      <pre>{log_html}</pre>
    </section>
  </main>
</body>
</html>
"""


def build_excel_zip(outdir: Path) -> bytes:
    buffer = io.BytesIO()
    written = 0
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for filename, _, kind in OUTPUT_FILES:
            if kind != "xlsx":
                continue
            path = outdir / filename
            if not path.is_file():
                continue
            archive.write(path, f"excel/{filename}")
            written += 1
    if written == 0:
        raise FileNotFoundError("no excel output")
    return buffer.getvalue()


def build_bundle_zip(job: dict[str, Any], outdir: Path) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("report.html", build_report_html(job))
        archive.writestr(
            "job_summary.json",
            json.dumps(job, ensure_ascii=False, indent=2, default=json_default),
        )
        for filename, _, _ in OUTPUT_FILES:
            path = outdir / filename
            if path.is_file():
                archive.write(path, f"outputs/{filename}")
    return buffer.getvalue()


DETAIL_HTML = r"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>BOM Check 明细</title>
  <style>
    :root {
      color-scheme: light;
      --bg: #f6f7f9;
      --panel: #ffffff;
      --ink: #16211f;
      --muted: #60706c;
      --line: #d8e0dd;
      --primary: #0b6b5f;
      --primary-dark: #074d44;
      --accent: #c96a28;
      --danger: #b42318;
      --warn: #a85a13;
      --ok: #11845b;
      --focus: #2563eb;
      font-family: Inter, "Segoe UI", "Microsoft YaHei", system-ui, sans-serif;
    }
    * { box-sizing: border-box; }
    body { margin: 0; min-width: 320px; background: var(--bg); color: var(--ink); line-height: 1.5; }
    button, input { font: inherit; }
    a:focus-visible, button:focus-visible { outline: 3px solid color-mix(in srgb, var(--focus) 45%, transparent); outline-offset: 2px; }
    .app { width: min(1740px, calc(100vw - 32px)); margin: 0 auto; padding: 18px 0 28px; }
    .topbar { display: flex; justify-content: space-between; align-items: flex-start; gap: 14px; margin-bottom: 14px; }
    h1 { margin: 0; font-size: 24px; line-height: 1.15; letter-spacing: 0; }
    .subline { margin-top: 4px; color: var(--muted); font-size: 13px; overflow-wrap: anywhere; }
    .actions { display: flex; flex-wrap: wrap; gap: 8px; justify-content: flex-end; }
    .btn { min-height: 36px; display: inline-flex; align-items: center; justify-content: center; padding: 7px 12px; border: 1px solid var(--line); border-radius: 6px; background: #fff; color: var(--primary); font-size: 13px; font-weight: 800; text-decoration: none; white-space: nowrap; cursor: pointer; }
    .btn:hover { border-color: var(--primary); background: #eef7f4; }
    .btn.primary { color: #fff; background: var(--primary); border-color: var(--primary); }
    .btn.primary:hover { background: var(--primary-dark); }
    .metrics { display: grid; grid-template-columns: repeat(5, minmax(120px, 1fr)); gap: 10px; margin-bottom: 14px; }
    .metric { min-height: 70px; padding: 10px 12px; border: 1px solid var(--line); border-radius: 8px; background: #fff; }
    .metric strong { display: block; font-size: 24px; line-height: 1.1; }
    .metric span { display: block; margin-top: 4px; color: var(--muted); font-size: 12px; font-weight: 800; }
    .layout { display: grid; grid-template-columns: minmax(360px, 470px) minmax(0, 1fr); gap: 14px; align-items: start; }
    .panel { border: 1px solid var(--line); border-radius: 8px; background: var(--panel); overflow: hidden; }
    .panel-head { min-height: 48px; display: flex; align-items: center; justify-content: space-between; gap: 10px; padding: 10px 12px; border-bottom: 1px solid var(--line); }
    .panel-title { margin: 0; font-size: 15px; font-weight: 900; }
    .search { width: 180px; min-height: 34px; border: 1px solid var(--line); border-radius: 6px; padding: 6px 9px; background: #fff; }
    .issue-list { max-height: calc(100vh - 230px); min-height: 520px; overflow: auto; }
    .issue { width: 100%; display: block; padding: 11px 12px; border: 0; border-bottom: 1px solid #e7ecea; border-left: 4px solid transparent; background: #fff; color: inherit; text-align: left; cursor: pointer; }
    .issue:hover { background: #f3f8f6; }
    .issue.selected { border-left-color: var(--accent); background: #fff7e8; }
    .issue.error .severity { color: var(--danger); }
    .issue.warning .severity { color: var(--warn); }
    .issue-title { display: flex; gap: 8px; align-items: baseline; justify-content: space-between; }
    .refdes { font-size: 16px; font-weight: 900; overflow-wrap: anywhere; }
    .severity { font-size: 12px; font-weight: 900; text-transform: uppercase; }
    .issue-meta { margin-top: 5px; color: var(--muted); font-size: 12px; display: flex; flex-wrap: wrap; gap: 8px; }
    .issue-desc { margin-top: 6px; color: #2c3936; font-size: 13px; overflow-wrap: anywhere; }
    .viewer-body { padding: 12px; }
    .locator-head { display: flex; align-items: center; justify-content: space-between; gap: 12px; margin-bottom: 10px; }
    .locator-title { font-weight: 900; overflow-wrap: anywhere; }
    .locator-meta { color: var(--muted); font-size: 13px; }
    .image-shell { display: grid; place-items: center; min-height: 600px; max-height: calc(100vh - 220px); overflow: auto; border: 1px solid var(--line); border-radius: 8px; background: #fdfefe; }
    .image-shell img { display: block; max-width: 100%; height: auto; }
    .empty { min-height: 280px; display: grid; place-items: center; padding: 24px; color: var(--muted); text-align: center; }
    .kv { display: grid; grid-template-columns: 96px 1fr; gap: 7px 10px; margin-top: 12px; color: #283633; font-size: 13px; }
    .kv dt { color: var(--muted); font-weight: 800; }
    .kv dd { margin: 0; overflow-wrap: anywhere; }
    .failed { color: var(--danger); font-weight: 800; }
    @media (max-width: 980px) {
      .app { width: min(100% - 20px, 1740px); }
      .topbar, .locator-head { display: grid; }
      .actions { justify-content: flex-start; }
      .metrics, .layout { grid-template-columns: 1fr; }
      .issue-list { max-height: none; min-height: 0; }
      .image-shell { min-height: 360px; max-height: none; }
    }
  </style>
</head>
<body>
  <main class="app">
    <header class="topbar">
      <div>
        <h1>BOM Check 明细</h1>
        <div id="subtitle" class="subline"></div>
      </div>
      <nav id="actions" class="actions"></nav>
    </header>
    <section id="metrics" class="metrics"></section>
    <section class="layout">
      <aside class="panel">
        <div class="panel-head">
          <h2 class="panel-title">异常列表</h2>
          <input id="search" class="search" type="search" placeholder="搜索">
        </div>
        <div id="issueList" class="issue-list"></div>
      </aside>
      <section class="panel">
        <div class="panel-head">
          <h2 class="panel-title">PDF 定位</h2>
          <a id="pdfButton" class="btn" href="#" target="_blank" rel="noopener">打开 PDF</a>
        </div>
        <div id="viewer" class="viewer-body"></div>
      </section>
    </section>
  </main>
  <script>
    const job = __JOB_JSON__;
    const issues = job.previews?.issues || {};
    const headers = issues.headers || [];
    const rows = issues.rows || [];
    const targets = issues.targets || [];
    const files = job.files || [];
    const pdfFile = files.find((file) => file.kind === "pdf");
    const searchInput = document.getElementById("search");
    const issueList = document.getElementById("issueList");
    const viewer = document.getElementById("viewer");
    const pdfButton = document.getElementById("pdfButton");
    let selectedIndex = Math.max(0, targets.findIndex((target) => target?.has_location));
    let filterText = "";

    function escapeHtml(value) {
      return String(value ?? "")
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;")
        .replaceAll("'", "&#039;");
    }

    function cell(row, names) {
      const normalized = names.map((name) => name.toLowerCase());
      const index = headers.findIndex((header) => normalized.includes(String(header).toLowerCase()));
      return index >= 0 ? row[index] : "";
    }

    function metric(label, value) {
      return `<div class="metric"><strong>${escapeHtml(value ?? 0)}</strong><span>${escapeHtml(label)}</span></div>`;
    }

    function actionLinks() {
      const id = encodeURIComponent(job.id);
      const links = [
        { label: "任务中心", href: "/", primary: false },
        { label: "打包下载", href: `/jobs/${id}/bundle.zip`, primary: true },
        { label: "下载页面", href: `/jobs/${id}/report.html?download=1`, primary: false },
        { label: "下载 Excel", href: `/jobs/${id}/excel.zip`, primary: false },
      ];
      if (pdfFile) {
        links.push({ label: "下载 PDF", href: pdfFile.download_url, primary: false });
      }
      document.getElementById("actions").innerHTML = links.map((link) =>
        `<a class="btn ${link.primary ? "primary" : ""}" href="${link.href}">${escapeHtml(link.label)}</a>`
      ).join("");
    }

    function renderHeader() {
      const summary = job.summary || {};
      const counts = summary.issue_counts || {};
      document.getElementById("subtitle").textContent = `${job.inputs?.bom || ""} · ${job.inputs?.pdf || ""} · ${job.message || job.status}`;
      document.getElementById("metrics").innerHTML = [
        metric("BOM 行", summary.bom_items),
        metric("PDF 位号", summary.pdf_refs),
        metric("异常", summary.issues),
        metric("错误", counts.error || 0),
        metric("警告", counts.warning || 0),
      ].join("");
      actionLinks();
    }

    function rowText(row, target) {
      return [
        target?.severity,
        target?.rule_id,
        target?.refdes,
        target?.title,
        ...(row || []),
      ].join(" ").toLowerCase();
    }

    function renderIssues() {
      const visible = rows
        .map((row, index) => ({ row, target: targets[index] || {}, index }))
        .filter((item) => rowText(item.row, item.target).includes(filterText));
      if (job.status !== "done") {
        issueList.innerHTML = `<div class="empty ${job.status === "failed" ? "failed" : ""}">${escapeHtml(job.message || "任务处理中")}</div>`;
        return;
      }
      if (!visible.length) {
        issueList.innerHTML = `<div class="empty">暂无异常</div>`;
        return;
      }
      issueList.innerHTML = visible.map(({ row, target, index }) => {
        const severity = target.severity || cell(row, ["Severity"]) || "";
        const title = target.title || cell(row, ["Title"]) || "";
        const refdes = target.refdes || cell(row, ["RefDes"]) || "";
        const rule = target.rule_id || cell(row, ["Rule_ID"]) || "";
        const bomRow = target.bom_row || cell(row, ["BOM_Row"]) || "";
        const page = target.page || target.pdf_page || cell(row, ["PDF_Page"]) || "";
        const selected = index === selectedIndex;
        const locateText = target.has_location ? `第 ${target.page} 页` : "无坐标";
        return `
          <button class="issue ${escapeHtml(severity)} ${selected ? "selected" : ""}" type="button" data-index="${index}">
            <span class="issue-title">
              <span class="refdes">${escapeHtml(refdes || "-")}</span>
              <span class="severity">${escapeHtml(severity || "-")}</span>
            </span>
            <span class="issue-meta">
              <span>${escapeHtml(rule || "-")}</span>
              <span>BOM ${escapeHtml(bomRow || "-")}</span>
              <span>${escapeHtml(page ? "PDF " + page : locateText)}</span>
              <span>${escapeHtml(locateText)}</span>
            </span>
            <span class="issue-desc">${escapeHtml(title || "-")}</span>
          </button>
        `;
      }).join("");
      issueList.querySelectorAll("[data-index]").forEach((button) => {
        button.addEventListener("click", () => {
          selectedIndex = Number(button.dataset.index);
          renderIssues();
          renderViewer();
        });
      });
    }

    function renderViewer() {
      if (job.status !== "done") {
        viewer.innerHTML = `<div class="empty ${job.status === "failed" ? "failed" : ""}">${escapeHtml(job.message || "任务处理中")}</div>`;
        pdfButton.style.display = "none";
        return;
      }
      const target = targets[selectedIndex] || {};
      if (!pdfFile) {
        viewer.innerHTML = `<div class="empty">暂无标注 PDF</div>`;
        pdfButton.style.display = "none";
        return;
      }
      pdfButton.style.display = "";
      pdfButton.href = target.has_location && target.page
        ? `${pdfFile.url}#page=${target.page}&view=FitH`
        : pdfFile.url;
      if (!target.has_location) {
        viewer.innerHTML = `
          <div class="empty">该异常没有 PDF 坐标</div>
          <dl class="kv">
            <dt>位号</dt><dd>${escapeHtml(target.refdes || "")}</dd>
            <dt>规则</dt><dd>${escapeHtml(target.rule_id || "")}</dd>
            <dt>说明</dt><dd>${escapeHtml(target.title || "")}</dd>
          </dl>
        `;
        return;
      }
      const bbox = encodeURIComponent(target.bbox.join(","));
      const src = `/jobs/${encodeURIComponent(job.id)}/locate.png?page_index=${target.page_index}&bbox=${bbox}&t=${Date.now()}`;
      viewer.innerHTML = `
        <div class="locator-head">
          <div>
            <div class="locator-title">${escapeHtml(target.refdes || "")} · 第 ${escapeHtml(target.page)} 页</div>
            <div class="locator-meta">${escapeHtml(target.rule_id || "")} · ${escapeHtml(target.title || "")}</div>
          </div>
          <a class="btn" href="${pdfButton.href}" target="_blank" rel="noopener">PDF 该页</a>
        </div>
        <div class="image-shell"><img src="${src}" alt="异常定位高亮预览"></div>
        <dl class="kv">
          <dt>上下文</dt><dd>${escapeHtml(target.context || target.raw_text || "")}</dd>
          <dt>BOM 行</dt><dd>${escapeHtml(target.bom_row || "")}</dd>
          <dt>坐标</dt><dd>${escapeHtml(target.bbox.join(", "))}</dd>
        </dl>
      `;
    }

    searchInput.addEventListener("input", () => {
      filterText = searchInput.value.trim().toLowerCase();
      renderIssues();
    });

    renderHeader();
    renderIssues();
    renderViewer();
  </script>
</body>
</html>
"""


def build_detail_html(job: dict[str, Any]) -> str:
    return DETAIL_HTML.replace("__JOB_JSON__", json_for_script(job))


APP_HTML = r"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>BOM Check</title>
  <style>
    :root {
      color-scheme: light;
      --bg: #f6f7f9;
      --panel: #ffffff;
      --ink: #16211f;
      --muted: #60706c;
      --line: #d8e0dd;
      --primary: #0b6b5f;
      --primary-dark: #074d44;
      --accent: #c96a28;
      --danger: #b42318;
      --warn: #a85a13;
      --ok: #11845b;
      --focus: #2563eb;
      font-family: Inter, "Segoe UI", "Microsoft YaHei", system-ui, sans-serif;
    }
    * { box-sizing: border-box; }
    body { margin: 0; min-width: 320px; background: var(--bg); color: var(--ink); line-height: 1.5; }
    button, input { font: inherit; }
    button:focus-visible, input:focus-visible, a:focus-visible { outline: 3px solid color-mix(in srgb, var(--focus) 45%, transparent); outline-offset: 2px; }
    .app { width: min(1320px, calc(100vw - 32px)); margin: 0 auto; padding: 22px 0 32px; }
    .topbar { display: flex; align-items: flex-start; justify-content: space-between; gap: 16px; margin-bottom: 16px; }
    h1 { margin: 0; font-size: 28px; line-height: 1.15; letter-spacing: 0; }
    .subtitle { margin-top: 4px; color: var(--muted); font-size: 13px; }
    .top-actions, .form-actions, .job-actions { display: flex; align-items: center; flex-wrap: wrap; gap: 8px; }
    .btn, button.btn { min-height: 38px; display: inline-flex; align-items: center; justify-content: center; padding: 7px 12px; border: 1px solid var(--line); border-radius: 6px; background: #fff; color: var(--primary); font-size: 13px; font-weight: 800; text-decoration: none; white-space: nowrap; cursor: pointer; }
    .btn:hover { border-color: var(--primary); background: #eef7f4; }
    .btn.primary { color: #fff; background: var(--primary); border-color: var(--primary); }
    .btn.primary:hover { background: var(--primary-dark); }
    .btn:disabled { color: #fff; background: #98aaa5; border-color: #98aaa5; cursor: not-allowed; }
    .grid { display: grid; grid-template-columns: minmax(320px, 410px) minmax(0, 1fr); gap: 16px; align-items: start; }
    .panel { border: 1px solid var(--line); border-radius: 8px; background: var(--panel); overflow: hidden; }
    .panel-head { min-height: 48px; display: flex; align-items: center; justify-content: space-between; gap: 10px; padding: 10px 14px; border-bottom: 1px solid var(--line); }
    .panel-title { margin: 0; font-size: 15px; font-weight: 900; }
    .upload { padding: 14px; }
    .field { display: grid; gap: 7px; margin-bottom: 13px; }
    label { color: var(--muted); font-size: 13px; font-weight: 800; }
    input[type="file"] { width: 100%; min-height: 44px; padding: 9px; border: 1px solid var(--line); border-radius: 6px; background: #fff; color: var(--ink); cursor: pointer; }
    .form-actions { margin-top: 4px; }
    .form-actions .primary { flex: 1 1 180px; }
    .progress { height: 8px; margin-top: 13px; overflow: hidden; border-radius: 999px; background: #e7ecea; }
    .progress span { display: block; width: 0; height: 100%; background: var(--primary); }
    .notice { min-height: 22px; margin-top: 10px; color: var(--muted); font-size: 13px; overflow-wrap: anywhere; }
    .notice.error { color: var(--danger); font-weight: 800; }
    .queue { display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 10px; padding: 12px; }
    .queue-item { min-height: 72px; padding: 10px; border: 1px solid #e1e8e5; border-radius: 6px; background: #fbfcfc; }
    .queue-item strong { display: block; font-size: 24px; line-height: 1.1; }
    .queue-item span { display: block; margin-top: 4px; color: var(--muted); font-size: 12px; font-weight: 800; }
    .jobs { display: grid; gap: 10px; padding: 12px; }
    .job { border: 1px solid #e1e8e5; border-radius: 8px; background: #fff; overflow: hidden; }
    .job-main { display: grid; grid-template-columns: minmax(0, 1fr) auto; gap: 12px; padding: 12px; align-items: start; }
    .job-name { margin: 0; font-size: 15px; font-weight: 900; overflow-wrap: anywhere; }
    .job-meta { display: flex; flex-wrap: wrap; gap: 8px; margin-top: 5px; color: var(--muted); font-size: 12px; }
    .status { min-height: 28px; display: inline-flex; align-items: center; padding: 4px 9px; border: 1px solid var(--line); border-radius: 999px; font-size: 12px; font-weight: 900; white-space: nowrap; }
    .status.queued, .status.running { color: var(--warn); border-color: #e9c49d; background: #fff8ef; }
    .status.done { color: var(--ok); border-color: #a7d8c4; background: #effaf5; }
    .status.failed { color: var(--danger); border-color: #f0b8b2; background: #fff3f1; }
    .job-stats { display: grid; grid-template-columns: repeat(4, minmax(86px, 1fr)); gap: 8px; padding: 0 12px 12px; }
    .stat { padding: 8px; border: 1px solid #e7ecea; border-radius: 6px; background: #fbfcfc; }
    .stat strong { display: block; font-size: 18px; line-height: 1.1; }
    .stat span { display: block; margin-top: 3px; color: var(--muted); font-size: 11px; font-weight: 800; }
    .job-actions { padding: 10px 12px; border-top: 1px solid #e7ecea; background: #fbfcfc; }
    .job-actions .primary { background: var(--primary); color: #fff; border-color: var(--primary); }
    .empty { min-height: 180px; display: grid; place-items: center; padding: 24px; color: var(--muted); text-align: center; }
    @media (max-width: 920px) {
      .app { width: min(100% - 20px, 1320px); }
      .topbar, .job-main { display: grid; }
      .grid, .queue, .job-stats { grid-template-columns: 1fr; }
      .top-actions { justify-content: flex-start; }
    }
  </style>
</head>
<body>
  <main class="app">
    <header class="topbar">
      <div>
        <h1>BOM Check</h1>
        <div class="subtitle">任务中心</div>
      </div>
      <div class="top-actions">
        <a class="btn" href="/" target="_blank" rel="noopener">新窗口</a>
        <button id="refreshButton" class="btn" type="button">刷新</button>
      </div>
    </header>
    <div class="grid">
      <section class="panel">
        <div class="panel-head">
          <h2 class="panel-title">新建任务</h2>
        </div>
        <form id="uploadForm" class="upload">
          <div class="field">
            <label for="bom">BOM Excel</label>
            <input id="bom" name="bom" type="file" accept=".xlsx,.xlsm,.xltx,.xltm,.csv" required>
          </div>
          <div class="field">
            <label for="pdf">原理图 PDF</label>
            <input id="pdf" name="pdf" type="file" accept=".pdf" required>
          </div>
          <div class="form-actions">
            <button id="runButton" class="btn primary" type="submit">运行检查</button>
          </div>
          <div class="progress" aria-hidden="true"><span id="progressBar"></span></div>
          <div id="notice" class="notice" role="status"></div>
        </form>
      </section>
      <section class="panel">
        <div class="panel-head">
          <h2 class="panel-title">队列</h2>
          <span id="healthTime" class="subtitle"></span>
        </div>
        <div id="queue" class="queue"></div>
      </section>
    </div>
    <section class="panel" style="margin-top:16px">
      <div class="panel-head">
        <h2 class="panel-title">最近任务</h2>
        <span id="jobCount" class="subtitle"></span>
      </div>
      <div id="jobs" class="jobs"></div>
    </section>
  </main>
  <script>
    const form = document.getElementById("uploadForm");
    const runButton = document.getElementById("runButton");
    const refreshButton = document.getElementById("refreshButton");
    const progressBar = document.getElementById("progressBar");
    const notice = document.getElementById("notice");
    const queuePanel = document.getElementById("queue");
    const jobsPanel = document.getElementById("jobs");
    const healthTime = document.getElementById("healthTime");
    const jobCount = document.getElementById("jobCount");

    function escapeHtml(value) {
      return String(value ?? "")
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;")
        .replaceAll("'", "&#039;");
    }

    function setNotice(text, isError = false) {
      notice.textContent = text || "";
      notice.className = "notice" + (isError ? " error" : "");
    }

    function queueMetric(label, value) {
      return `<div class="queue-item"><strong>${escapeHtml(value)}</strong><span>${escapeHtml(label)}</span></div>`;
    }

    function statusLabel(job) {
      if (job.status === "queued") return job.message || "排队中";
      if (job.status === "running") return "运行中";
      if (job.status === "done") return "完成";
      if (job.status === "failed") return "失败";
      return job.status || "-";
    }

    function stat(label, value) {
      return `<div class="stat"><strong>${escapeHtml(value ?? 0)}</strong><span>${escapeHtml(label)}</span></div>`;
    }

    function actionHtml(job) {
      if (job.status !== "done") return "";
      const id = encodeURIComponent(job.id);
      const pdf = (job.files || []).find((file) => file.kind === "pdf");
      const links = [
        { label: "查看明细", href: `/jobs/${id}/detail`, cls: "primary", target: "_blank" },
        { label: "打包下载", href: `/jobs/${id}/bundle.zip`, cls: "" },
        { label: "下载页面", href: `/jobs/${id}/report.html?download=1`, cls: "" },
        { label: "下载 Excel", href: `/jobs/${id}/excel.zip`, cls: "" },
      ];
      if (pdf) links.push({ label: "下载 PDF", href: pdf.download_url, cls: "" });
      return `<div class="job-actions">${links.map((link) =>
        `<a class="btn ${link.cls}" href="${link.href}" ${link.target ? `target="${link.target}" rel="noopener"` : ""}>${escapeHtml(link.label)}</a>`
      ).join("")}</div>`;
    }

    function renderJob(job) {
      const summary = job.summary || {};
      const counts = summary.issue_counts || {};
      const title = job.inputs?.bom || job.id;
      return `
        <article class="job">
          <div class="job-main">
            <div>
              <h3 class="job-name">${escapeHtml(title)}</h3>
              <div class="job-meta">
                <span>${escapeHtml(job.inputs?.pdf || "")}</span>
                <span>${escapeHtml(job.created_at || "")}</span>
                <span>${escapeHtml(job.id || "")}</span>
              </div>
            </div>
            <span class="status ${escapeHtml(job.status || "")}">${escapeHtml(statusLabel(job))}</span>
          </div>
          ${job.status === "done" ? `
            <div class="job-stats">
              ${stat("BOM 行", summary.bom_items)}
              ${stat("PDF 位号", summary.pdf_refs)}
              ${stat("错误", counts.error || 0)}
              ${stat("警告", counts.warning || 0)}
            </div>
          ` : ""}
          ${job.status === "failed" ? `<div class="notice error" style="padding:0 12px 12px">${escapeHtml(job.message || "失败")}</div>` : ""}
          ${actionHtml(job)}
        </article>
      `;
    }

    function renderQueue(health, jobs) {
      const queued = jobs.filter((job) => job.status === "queued").length;
      const running = jobs.filter((job) => job.status === "running").length;
      queuePanel.innerHTML = [
        queueMetric("运行中", health?.running ?? running),
        queueMetric("排队中", health?.queued ?? queued),
        queueMetric("Worker", health?.workers ?? 1),
      ].join("");
      healthTime.textContent = health?.time || "";
    }

    async function refreshAll() {
      const [healthResponse, jobsResponse] = await Promise.all([
        fetch("/health", { cache: "no-store" }),
        fetch("/api/jobs", { cache: "no-store" }),
      ]);
      const health = await healthResponse.json();
      const jobData = await jobsResponse.json();
      if (!healthResponse.ok) throw new Error(health.error || "读取队列失败");
      if (!jobsResponse.ok) throw new Error(jobData.error || "读取任务失败");
      const jobs = jobData.jobs || [];
      renderQueue(health, jobs);
      jobCount.textContent = jobs.length ? `${jobs.length} 个` : "";
      jobsPanel.innerHTML = jobs.length ? jobs.map(renderJob).join("") : `<div class="empty">暂无任务</div>`;
    }

    form.addEventListener("submit", (event) => {
      event.preventDefault();
      setNotice("");
      runButton.disabled = true;
      progressBar.style.width = "0";
      const data = new FormData(form);
      const xhr = new XMLHttpRequest();
      xhr.open("POST", "/api/jobs");
      xhr.upload.onprogress = (event) => {
        if (event.lengthComputable) {
          progressBar.style.width = `${Math.round((event.loaded / event.total) * 100)}%`;
        }
      };
      xhr.onload = () => {
        try {
          const response = JSON.parse(xhr.responseText);
          if (xhr.status < 200 || xhr.status >= 300) {
            throw new Error(response.error || "上传失败");
          }
          setNotice(response.job.message || "已加入队列");
          form.reset();
          refreshAll().catch((error) => setNotice(error.message, true));
        } catch (error) {
          setNotice(error.message || String(error), true);
        } finally {
          runButton.disabled = false;
          progressBar.style.width = "0";
        }
      };
      xhr.onerror = () => {
        setNotice("网络连接失败", true);
        runButton.disabled = false;
        progressBar.style.width = "0";
      };
      xhr.send(data);
    });

    refreshButton.addEventListener("click", () => refreshAll().catch((error) => setNotice(error.message, true)));
    refreshAll().catch((error) => setNotice(error.message, true));
    setInterval(() => refreshAll().catch(() => {}), 2000);
  </script>
</body>
</html>
"""


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(prog="bomcheck-web")
    parser.add_argument("--host", default="0.0.0.0", help="HTTP listen host")
    parser.add_argument("--port", default=8088, type=int, help="HTTP listen port")
    parser.add_argument("--jobs-dir", default="var/jobs", help="directory for uploaded files and outputs")
    parser.add_argument("--max-upload-mb", default=300, type=int, help="max total request size")
    parser.add_argument("--workers", default=1, type=int, help="number of background check workers")
    args = parser.parse_args(argv)

    jobs_root = Path(args.jobs_dir).resolve()
    server = BomCheckServer(
        (args.host, args.port),
        jobs_root=jobs_root,
        max_upload_bytes=args.max_upload_mb * 1024 * 1024,
        worker_count=args.workers,
    )
    url_host = "127.0.0.1" if args.host in {"0.0.0.0", "::"} else args.host
    print(f"BOM Check Web running: http://{url_host}:{args.port}")
    print(f"Jobs directory: {jobs_root}")
    print(f"Workers: {server.worker_count}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopping BOM Check Web")
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
